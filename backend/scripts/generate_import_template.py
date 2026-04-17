#!/usr/bin/env python
"""Generate the 4-sheet example import workbook at docs/example_import.xlsx.

Sheet layout matches `app.services.imports` expectations. Frame → fixation
linkage is by a human-readable `frame_ref` column (Excel row reference).
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook


def _write(path: Path) -> None:
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    if default is not None:
        wb.remove(default)

    frames = wb.create_sheet("physical_frames")
    frames.append(
        [
            "frame_ref",
            "commodity",
            "side",
            "quantity_tons",
            "delivery_start",
            "delivery_end",
            "counterparty",
            "notes",
        ]
    )
    frames.append(
        ["F1", "soja", "sell", 1000, "2026-05-01", "2026-07-31", "Cargill BR", "safra 25/26"]
    )
    frames.append(["F2", "milho", "buy", 500, "2026-06-01", "2026-08-31", "ADM", "proxy ZC"])
    frames.append(["F3", "soja", "sell", 2000, "2026-09-01", "2026-12-31", "COFCO", "safra nova"])

    fixations = wb.create_sheet("physical_fixations")
    fixations.append(
        [
            "frame_ref",
            "fixation_mode",
            "quantity_tons",
            "fixation_date",
            "cbot_fixed",
            "basis_fixed",
            "fx_fixed",
            "reference_cbot_contract",
            "notes",
        ]
    )
    fixations.append(
        ["F1", "flat", 300, "2026-04-10", 1420.25, 0.50, 5.00, "ZSK26", "fixacao parcial"]
    )
    fixations.append(
        ["F1", "cbot", 200, "2026-04-15", 1415.75, None, None, "ZSK26", "bolsa apenas"]
    )
    fixations.append(["F3", "basis", 500, "2026-04-12", None, 0.60, None, "ZSX26", "fixou premio"])

    cbot = wb.create_sheet("cbot")
    cbot.append(
        [
            "commodity",
            "instrument",
            "side",
            "contract",
            "quantity_contracts",
            "trade_date",
            "trade_price",
            "maturity_date",
            "option_type",
            "strike",
            "barrier_type",
            "barrier_level",
            "rebate",
            "counterparty",
            "notes",
        ]
    )
    cbot.append(
        [
            "soja",
            "future",
            "sell",
            "ZSK26",
            5,
            "2026-04-10",
            1420.0,
            "2026-05-14",
            None,
            None,
            None,
            None,
            None,
            "broker",
            "hedge",
        ]
    )
    cbot.append(
        [
            "soja",
            "european_option",
            "buy",
            "ZSK26",
            2,
            "2026-04-15",
            25.5,
            "2026-05-14",
            "call",
            1450.0,
            None,
            None,
            None,
            "broker",
            "cap de preco",
        ]
    )

    basis = wb.create_sheet("basis")
    basis.append(
        [
            "commodity",
            "side",
            "quantity_tons",
            "trade_date",
            "basis_price",
            "delivery_date",
            "reference_cbot_contract",
            "counterparty",
            "notes",
        ]
    )
    basis.append(
        ["soja", "buy", 500, "2026-04-10", 0.55, "2026-07-15", "ZSK26", "Bunge", "fwd premio"]
    )
    basis.append(
        ["milho", "sell", 250, "2026-04-12", -0.35, "2026-08-15", "ZCN26", "Cargill", "curto"]
    )

    fx = wb.create_sheet("fx")
    fx.append(
        [
            "instrument",
            "side",
            "notional_usd",
            "trade_date",
            "trade_rate",
            "maturity_date",
            "option_type",
            "strike",
            "barrier_type",
            "barrier_level",
            "rebate",
            "counterparty",
            "notes",
        ]
    )
    fx.append(
        [
            "ndf",
            "sell",
            500000,
            "2026-04-10",
            5.0234,
            "2026-07-15",
            None,
            None,
            None,
            None,
            None,
            "Itau",
            "hedge usd",
        ]
    )
    fx.append(
        [
            "european_option",
            "buy",
            1000000,
            "2026-04-15",
            5.10,
            "2026-07-15",
            "call",
            5.20,
            None,
            None,
            None,
            "BTG",
            "cap cambial",
        ]
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate docs/example_import.xlsx")
    parser.add_argument(
        "--output",
        default=str(Path(__file__).resolve().parents[2] / "docs" / "example_import.xlsx"),
        help="Output path (default: docs/example_import.xlsx at repo root)",
    )
    args = parser.parse_args()
    out = Path(args.output)
    _write(out)
    print(f"wrote {out} ({out.stat().st_size} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

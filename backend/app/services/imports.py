"""Multi-sheet Excel/CSV import for positions.

Sheet names recognized (case-insensitive): physical_frames, physical_fixations,
cbot, basis, fx. Portuguese column aliases are normalized to the English
schema field names before Pydantic validation. Each row that fails validation
is collected as a `RowError`; the rest are parsed into `*In` schemas.

`parse_workbook` is pure (no DB). The caller decides whether to preview or
commit.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError

from app.schemas.basis import BasisForwardIn
from app.schemas.cbot import CBOTDerivativeIn
from app.schemas.fx import FXDerivativeIn
from app.schemas.physical import PhysicalFixationIn, PhysicalFrameIn

# --------------------------------------------------------------------------
# Column aliasing: Portuguese / shorthand → canonical schema field name.
# Applied per-sheet to normalize headers before model validation.
# --------------------------------------------------------------------------

_COMMON_ALIASES: dict[str, str] = {
    "commodity": "commodity",
    "mercadoria": "commodity",
    "produto": "commodity",
    "side": "side",
    "lado": "side",
    "operacao": "side",
    "tons": "quantity_tons",
    "toneladas": "quantity_tons",
    "quantidade_tons": "quantity_tons",
    "quantity_tons": "quantity_tons",
    "contraparte": "counterparty",
    "counterparty": "counterparty",
    "notes": "notes",
    "observacoes": "notes",
    "notas": "notes",
}

_FRAME_ALIASES: dict[str, str] = {
    **_COMMON_ALIASES,
    "delivery_start": "delivery_start",
    "entrega_inicio": "delivery_start",
    "inicio_entrega": "delivery_start",
    "delivery_end": "delivery_end",
    "entrega_fim": "delivery_end",
    "fim_entrega": "delivery_end",
    "vencimento": "delivery_end",
    "frame_ref": "frame_ref",
}

_FIXATION_ALIASES: dict[str, str] = {
    "frame_ref": "frame_ref",
    "fixation_mode": "fixation_mode",
    "modo": "fixation_mode",
    "modalidade": "fixation_mode",
    "fixation_date": "fixation_date",
    "data_fixacao": "fixation_date",
    "tons": "quantity_tons",
    "toneladas": "quantity_tons",
    "quantity_tons": "quantity_tons",
    "cbot_fixed": "cbot_fixed",
    "cbot": "cbot_fixed",
    "bolsa": "cbot_fixed",
    "basis_fixed": "basis_fixed",
    "basis": "basis_fixed",
    "premio": "basis_fixed",
    "prêmio": "basis_fixed",
    "fx_fixed": "fx_fixed",
    "dolar": "fx_fixed",
    "dólar": "fx_fixed",
    "reference_cbot_contract": "reference_cbot_contract",
    "contrato_cbot": "reference_cbot_contract",
    "notes": "notes",
    "notas": "notes",
}

_CBOT_ALIASES: dict[str, str] = {
    **_COMMON_ALIASES,
    "instrument": "instrument",
    "instrumento": "instrument",
    "contract": "contract",
    "contrato": "contract",
    "quantity_contracts": "quantity_contracts",
    "contratos": "quantity_contracts",
    "trade_date": "trade_date",
    "data_operacao": "trade_date",
    "trade_price": "trade_price",
    "preco": "trade_price",
    "preço": "trade_price",
    "maturity_date": "maturity_date",
    "vencimento": "maturity_date",
    "option_type": "option_type",
    "tipo_opcao": "option_type",
    "strike": "strike",
    "barrier_type": "barrier_type",
    "barrier_level": "barrier_level",
    "barreira": "barrier_level",
    "rebate": "rebate",
}

_BASIS_ALIASES: dict[str, str] = {
    **_COMMON_ALIASES,
    "trade_date": "trade_date",
    "basis_price": "basis_price",
    "premio": "basis_price",
    "prêmio": "basis_price",
    "delivery_date": "delivery_date",
    "entrega": "delivery_date",
    "reference_cbot_contract": "reference_cbot_contract",
    "contrato_cbot": "reference_cbot_contract",
}

_FX_ALIASES: dict[str, str] = {
    "instrument": "instrument",
    "instrumento": "instrument",
    "side": "side",
    "lado": "side",
    "notional_usd": "notional_usd",
    "notional": "notional_usd",
    "trade_date": "trade_date",
    "trade_rate": "trade_rate",
    "taxa": "trade_rate",
    "maturity_date": "maturity_date",
    "vencimento": "maturity_date",
    "option_type": "option_type",
    "strike": "strike",
    "barrier_type": "barrier_type",
    "barrier_level": "barrier_level",
    "rebate": "rebate",
    "counterparty": "counterparty",
    "contraparte": "counterparty",
    "notes": "notes",
}


# --------------------------------------------------------------------------
# Data containers
# --------------------------------------------------------------------------


@dataclass
class RowError:
    sheet: str
    row_index: int  # 1-based to match Excel row numbers (header = row 1)
    errors: list[dict[str, Any]]


@dataclass
class ParsedFixation:
    frame_ref: str | None
    fixation: PhysicalFixationIn


@dataclass
class ImportPayload:
    frames: list[tuple[str | None, PhysicalFrameIn]] = field(default_factory=list)
    fixations: list[ParsedFixation] = field(default_factory=list)
    cbot: list[CBOTDerivativeIn] = field(default_factory=list)
    basis: list[BasisForwardIn] = field(default_factory=list)
    fx: list[FXDerivativeIn] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)

    @property
    def valid_count(self) -> int:
        return (
            len(self.frames) + len(self.fixations) + len(self.cbot) + len(self.basis) + len(self.fx)
        )

    @property
    def invalid_count(self) -> int:
        return len(self.errors)


# --------------------------------------------------------------------------
# Normalization helpers
# --------------------------------------------------------------------------


_SIDE_LOOKUP: dict[str, str] = {
    "buy": "buy",
    "sell": "sell",
    "compra": "buy",
    "venda": "sell",
    "long": "buy",
    "short": "sell",
    "c": "buy",
    "v": "sell",
}


def _normalize_header(raw: Any) -> str:
    if raw is None:
        return ""
    return str(raw).strip().lower().replace(" ", "_")


def _normalize_side(value: Any) -> Any:
    if isinstance(value, str):
        key = value.strip().lower()
        return _SIDE_LOOKUP.get(key, value)
    return value


def _map_row(raw: dict[str, Any], aliases: dict[str, str]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for key, val in raw.items():
        canon = aliases.get(_normalize_header(key))
        if canon is None:
            continue
        if canon == "side":
            val = _normalize_side(val)
        if isinstance(val, str):
            val = val.strip()
            if val == "":
                val = None
        mapped[canon] = val
    return mapped


def _validate(
    model: type[BaseModel], row: dict[str, Any], sheet: str, row_index: int
) -> BaseModel | RowError:
    try:
        return model.model_validate(row)
    except ValidationError as exc:
        return RowError(
            sheet=sheet,
            row_index=row_index,
            errors=[dict(e) for e in exc.errors()],
        )


def _iter_rows(sheet_obj: Any) -> list[dict[str, Any]]:
    rows = list(sheet_obj.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        out.append(dict(zip(headers, row, strict=False)))
    return out


# --------------------------------------------------------------------------
# Public entrypoint
# --------------------------------------------------------------------------


_SHEET_MAP = {
    "physical_frames": "frames",
    "physical_fixations": "fixations",
    "cbot": "cbot",
    "basis": "basis",
    "fx": "fx",
}


def parse_workbook(data: bytes) -> ImportPayload:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    payload = ImportPayload()
    normalized_sheets = {name.lower(): name for name in wb.sheetnames}

    for canon_name, kind in _SHEET_MAP.items():
        raw_name = normalized_sheets.get(canon_name)
        if raw_name is None:
            continue
        sheet_obj = wb[raw_name]
        rows = _iter_rows(sheet_obj)

        for idx, raw_row in enumerate(rows, start=2):  # row 2 = first data row in Excel
            if kind == "frames":
                frame_ref = raw_row.get("frame_ref") or raw_row.get("ref")
                mapped = _map_row(raw_row, _FRAME_ALIASES)
                mapped.pop("frame_ref", None)
                res = _validate(PhysicalFrameIn, mapped, canon_name, idx)
                if isinstance(res, RowError):
                    payload.errors.append(res)
                else:
                    assert isinstance(res, PhysicalFrameIn)
                    payload.frames.append((str(frame_ref) if frame_ref else None, res))
            elif kind == "fixations":
                mapped = _map_row(raw_row, _FIXATION_ALIASES)
                frame_ref = mapped.pop("frame_ref", None)
                res = _validate(PhysicalFixationIn, mapped, canon_name, idx)
                if isinstance(res, RowError):
                    payload.errors.append(res)
                else:
                    assert isinstance(res, PhysicalFixationIn)
                    payload.fixations.append(
                        ParsedFixation(
                            frame_ref=str(frame_ref) if frame_ref else None, fixation=res
                        )
                    )
            elif kind == "cbot":
                mapped = _map_row(raw_row, _CBOT_ALIASES)
                res = _validate(CBOTDerivativeIn, mapped, canon_name, idx)
                if isinstance(res, RowError):
                    payload.errors.append(res)
                else:
                    assert isinstance(res, CBOTDerivativeIn)
                    payload.cbot.append(res)
            elif kind == "basis":
                mapped = _map_row(raw_row, _BASIS_ALIASES)
                res = _validate(BasisForwardIn, mapped, canon_name, idx)
                if isinstance(res, RowError):
                    payload.errors.append(res)
                else:
                    assert isinstance(res, BasisForwardIn)
                    payload.basis.append(res)
            elif kind == "fx":
                mapped = _map_row(raw_row, _FX_ALIASES)
                res = _validate(FXDerivativeIn, mapped, canon_name, idx)
                if isinstance(res, RowError):
                    payload.errors.append(res)
                else:
                    assert isinstance(res, FXDerivativeIn)
                    payload.fx.append(res)

    return payload
